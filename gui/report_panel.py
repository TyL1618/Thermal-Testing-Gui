"""
report_panel.py  ─  GOTECH HV3000 報告頁面

功能：
  1. 左側：選擇測試記錄（列表）+ 詳細資訊預覽
  2. 中側：匯入測試記錄至報告產生器 / 匯入至檢視器
  3. 右側（產生報告）：
       - 基本資訊欄位填寫（測試名稱、客戶名稱、使用介質等）
       - 即時折線圖預覽（time-mm-°C / °C-mm 切換）
       - 測試資料表格
       - 匯出報告（CSV）
  4. 移除測試資料 / 匯出測試記錄
"""
from __future__ import annotations
import os
import csv
import json
import time
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from collections import deque

from PyQt6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QGridLayout,
    QLabel, QPushButton, QGroupBox, QCheckBox,
    QComboBox, QLineEdit, QListWidget, QListWidgetItem,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QFrame, QSizePolicy, QMessageBox, QFileDialog,
    QSplitter, QScrollArea, QAbstractItemView,
    QRadioButton, QButtonGroup, QTextEdit,
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer, QRegularExpression
from PyQt6.QtGui import QFont, QColor, QRegularExpressionValidator
import pyqtgraph as pg

from core.machine import GotechMachine

# ── 色彩
C = {
    'bg_deep':    '#080a0d',
    'bg_panel':   '#0e1117',
    'bg_card':    '#141820',
    'bg_card2':   '#1a1f2a',
    'border':     '#2a3345',
    'border_hi':  '#3d4f6a',
    'amber':      '#ffc233',
    'amber_dim':  '#7a5500',
    'green':      '#00f080',
    'green_dim':  '#004d26',
    'red':        '#ff6060',
    'red_dim':    '#7a0000',
    'blue':       '#5cd9ff',
    'blue_dim':   '#00405a',
    'text_hi':    '#f0f4ff',
    'text_mid':   '#b0bcd4',
    'text_lo':    '#6a7a96',
}

CH_COLORS = ['#ff6b6b', '#ffa94d', '#69db7c', '#4fc3f7', '#da77f2', '#f783ac']

# ── 測試資料儲存目錄
DATA_DIR = Path("test_data")
DATA_DIR.mkdir(exist_ok=True)


def _input_style() -> str:
    return f"""
        background:{C['bg_card2']};
        color:{C['text_hi']};
        border:1px solid {C['border_hi']};
        border-radius:4px;
        padding:4px 8px;
        font-family:'Consolas','Courier New',monospace;
        font-size:11px;
    """

def _btn_style(color: str, small=False) -> str:
    pad = "4px 10px" if small else "7px 16px"
    return f"""
        QPushButton {{
            background: transparent;
            color: {color};
            border: 1px solid {color};
            border-radius: 4px;
            padding: {pad};
            font-family: 'Consolas','Courier New',monospace;
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 1px;
        }}
        QPushButton:hover {{ background: {color}33; }}
        QPushButton:pressed {{ background: {color}; color: #000; }}
        QPushButton:disabled {{ color:{C['text_lo']}; border-color:{C['border']}; }}
    """

# ── 禁止輸入 CSV 危險字元（逗號、雙引號、換行）的 QLineEdit
_SAFE_RE = QRegularExpression(r"[^,\"\r\n]*")

def _safe_line_edit(placeholder: str = "--") -> QLineEdit:
    """建立一個禁止輸入 , \" \\n 的 QLineEdit"""
    le = QLineEdit(placeholder)
    le.setValidator(QRegularExpressionValidator(_SAFE_RE))
    le.setStyleSheet(_input_style())
    return le


def _label(text: str, color=None, size=10, bold=False, spacing=1) -> QLabel:
    c = color or C['text_lo']
    w = "700" if bold else "400"
    lbl = QLabel(text)
    lbl.setStyleSheet(f"""
        color:{c};
        font-family:'Consolas','Courier New',monospace;
        font-size:{size}px;font-weight:{w};
        letter-spacing:{spacing}px;
        background:transparent;
    """)
    return lbl


# ── 測試記錄資料結構
class TestRecord:
    def __init__(self):
        self.record_id: str = ""
        self.name: str = "--"
        self.test_name: str = "--"
        self.customer: str = "--"
        self.serial_no: str = "--"
        self.material: str = "--"
        self.test_date: str = datetime.now().strftime("%Y/%m/%d")
        self.test_time: str = "--"
        self.medium: str = "--"
        self.test_method: str = "--"
        self.group: str = "--"
        self.client_addr: str = "--"
        # 各通道資料列表
        self.channels: List[Dict[str, Any]] = []
        # 折線圖 time/mm/temp 歷史
        self.time_data: List[float] = []
        self.deflection_data: Dict[int, List[float]] = {i: [] for i in range(6)}
        self.temp_data: Dict[int, List[float]] = {i: [] for i in range(6)}

    def to_dict(self) -> dict:
        return self.__dict__

    @staticmethod
    def from_dict(d: dict) -> "TestRecord":
        r = TestRecord()
        for k, v in d.items():
            setattr(r, k, v)
        return r


class ReportPanel(QWidget):
    """報告頁面"""

    def __init__(self, machine: GotechMachine):
        super().__init__()
        self.machine = machine
        self.records: List[TestRecord] = []
        self._selected_record: Optional[TestRecord] = None
        self._report_record: Optional[TestRecord] = None

        self._setup_style()
        self._setup_ui()
        self._load_records_from_disk()

        # 讓 machine 的測試資料可以自動接收（monitor_panel 停止後呼叫 add_test_record）

    # ─────────────────────────────────────────────
    def _setup_style(self):
        self.setStyleSheet(f"""
            QWidget {{ background:{C['bg_deep']}; color:{C['text_hi']}; }}
            QGroupBox {{
                border:1px solid {C['border']};
                border-radius:6px;
                margin-top:16px;
                padding:8px;
                font-family:'Consolas','Courier New',monospace;
                font-size:10px;
                color:{C['text_mid']};
                letter-spacing:2px;
            }}
            QGroupBox::title {{
                subcontrol-origin:margin;
                left:10px;
                padding:0 6px;
                color:{C['amber']};
            }}
            QLabel {{ background:transparent; }}
            QTableWidget {{
                background:{C['bg_card2']};
                border:1px solid {C['border_hi']};
                border-radius:4px;
                color:{C['text_hi']};
                font-family:'Consolas','Courier New',monospace;
                font-size:10px;
                gridline-color:{C['border']};
            }}
            QTableWidget::item:selected {{
                background:{C['amber_dim']};
                color:{C['amber']};
            }}
            QHeaderView::section {{
                background:{C['bg_card']};
                color:{C['text_mid']};
                border:none;
                border-bottom:1px solid {C['border']};
                padding:4px;
                font-family:'Consolas','Courier New',monospace;
                font-size:9px;
                letter-spacing:1px;
            }}
            QListWidget {{
                background:{C['bg_card2']};
                border:1px solid {C['border_hi']};
                border-radius:4px;
                color:{C['text_hi']};
                font-family:'Consolas','Courier New',monospace;
                font-size:11px;
            }}
            QListWidget::item:selected {{
                background:{C['amber_dim']};
                color:{C['amber']};
            }}
            QRadioButton {{
                color:{C['text_mid']};
                font-size:10px;
                font-family:'Consolas','Courier New',monospace;
                background:transparent;
            }}
            QRadioButton::indicator {{ width:13px;height:13px; }}
            QLineEdit {{
                {_input_style()}
            }}
        """)

    # ─────────────────────────────────────────────
    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        # 頁標題
        title = _label("▤   REPORT", C['amber'], size=13, bold=True, spacing=4)
        root.addWidget(title)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"color:{C['border']};")
        root.addWidget(sep)

        # 主體：左 / 右
        body = QHBoxLayout()
        body.setSpacing(12)

        body.addWidget(self._make_left_panel(), stretch=42)
        body.addWidget(self._make_right_panel(), stretch=58)

        root.addLayout(body, stretch=1)

    # ─────────────────────────────────────────────
    def _make_left_panel(self) -> QWidget:
        frame = QWidget()
        frame.setStyleSheet(f"background:{C['bg_panel']};border:1px solid {C['border']};border-radius:8px;")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(8)

        lay.addWidget(_label("請選擇測試名稱", C['text_mid'], size=10, spacing=1))

        # 記錄選擇表格
        self.tbl_records = QTableWidget(0, 2)
        self.tbl_records.setHorizontalHeaderLabels(["ID", "測試名稱"])
        self.tbl_records.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tbl_records.setColumnWidth(0, 40)
        self.tbl_records.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_records.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_records.verticalHeader().setVisible(False)
        self.tbl_records.setFixedHeight(140)
        self.tbl_records.itemSelectionChanged.connect(self._on_record_selected)
        lay.addWidget(self.tbl_records)

        # 詳細資訊（右側小欄）
        info_grp = QGroupBox("測試詳細資訊")
        info_lay = QGridLayout(info_grp)
        info_lay.setSpacing(4)
        info_lay.setContentsMargins(8, 16, 8, 8)

        fields = [
            ("測試名稱", "lbl_i_test_name"),
            ("客戶名稱", "lbl_i_customer"),
            ("流水批號", "lbl_i_serial"),
            ("材料名稱", "lbl_i_material"),
            ("測試日期", "lbl_i_date"),
        ]
        for row, (field, attr) in enumerate(fields):
            info_lay.addWidget(_label(field, C['text_lo'], size=9), row, 0)
            lbl = _label("--", C['text_hi'], size=10)
            setattr(self, attr, lbl)
            info_lay.addWidget(lbl, row, 1)

        lay.addWidget(info_grp)

        # 報告標題欄位
        lay.addWidget(_label("報告標題：", C['text_lo'], size=9))
        self.le_report_title = _safe_line_edit("HDT/VICAT Test Report")
        lay.addWidget(self.le_report_title)

        # 列印拾頭
        self.chk_print_header = QCheckBox("列印拾頭")
        self.chk_print_header.setStyleSheet(f"""
            QCheckBox {{ color:{C['text_mid']};font-size:10px;background:transparent; }}
            QCheckBox::indicator {{
                width:13px;height:13px;border-radius:2px;
                border:1px solid {C['border_hi']};background:{C['bg_card2']};
            }}
            QCheckBox::indicator:checked {{
                background:{C['amber']};border-color:{C['amber']};
            }}
        """)
        lay.addWidget(self.chk_print_header)

        # 操作按鈕
        btn_to_report = QPushButton("匯入測試記錄至報告產生器 >>")
        btn_to_report.setStyleSheet(_btn_style(C['amber']))
        btn_to_report.clicked.connect(self._import_to_report)
        lay.addWidget(btn_to_report)

        btn_to_viewer = QPushButton("<< 匯入測試記錄至檢視器")
        btn_to_viewer.setStyleSheet(_btn_style(C['blue']))
        btn_to_viewer.clicked.connect(self._import_to_viewer)
        lay.addWidget(btn_to_viewer)

        lay.addStretch()

        # 底部按鈕
        bottom_row = QHBoxLayout()
        btn_remove = QPushButton("移除測試資料")
        btn_remove.setStyleSheet(_btn_style(C['red'], small=True))
        btn_remove.clicked.connect(self._remove_record)
        bottom_row.addWidget(btn_remove)

        btn_export_rec = QPushButton("匯出測試記錄")
        btn_export_rec.setStyleSheet(_btn_style(C['green'], small=True))
        btn_export_rec.clicked.connect(self._export_record)
        bottom_row.addWidget(btn_export_rec)

        lay.addLayout(bottom_row)

        # 詳細資料表（下半部）
        lay.addWidget(_label("測試通道資料", C['text_mid'], size=9))
        self.tbl_ch_data = QTableWidget(0, 7)
        self.tbl_ch_data.setHorizontalHeaderLabels([
            "客戶名稱", "材料名稱", "流水批號", "組別", "測試日期", "測試時間", "測試方法"
        ])
        for col in range(7):
            self.tbl_ch_data.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.ResizeToContents
            )
        self.tbl_ch_data.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_ch_data.verticalHeader().setVisible(False)
        lay.addWidget(self.tbl_ch_data, stretch=1)

        return frame

    # ─────────────────────────────────────────────
    def _make_right_panel(self) -> QWidget:
        frame = QWidget()
        frame.setStyleSheet(f"background:{C['bg_panel']};border:1px solid {C['border']};border-radius:8px;")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(8)

        lay.addWidget(_label("產生報告", C['amber'], size=12, bold=True, spacing=3))

        # ── 上半：基本資訊 + 折線圖（左右）
        upper = QHBoxLayout()
        upper.setSpacing(10)

        # 基本資訊
        info_col = QVBoxLayout()
        info_col.setSpacing(4)

        fields_r = [
            ("測試名稱", "le_r_test_name"),
            ("客戶名稱", "le_r_customer"),
            ("使用介質", "le_r_medium"),
            ("流水批號", "le_r_serial"),
            ("材料名稱", "le_r_material"),
            ("測試方法", "le_r_method"),
            ("測試日期", "le_r_date"),
            ("客戶地址", "le_r_addr"),
        ]
        for label_text, attr in fields_r:
            info_col.addWidget(_label(label_text, C['text_lo'], size=9))
            le = _safe_line_edit("--")
            setattr(self, attr, le)
            info_col.addWidget(le)

        info_col.addStretch()
        upper.addLayout(info_col, stretch=35)

        # 折線圖預覽
        chart_col = QVBoxLayout()
        chart_col.setSpacing(4)

        pg.setConfigOption('background', C['bg_deep'])
        pg.setConfigOption('foreground', C['text_lo'])
        self.report_plot = pg.PlotWidget()
        self.report_plot.setBackground(C['bg_deep'])
        self.report_plot.setLabel('left', 'mm', color=C['text_mid'], size='8pt')
        self.report_plot.setLabel('bottom', '', color=C['text_mid'], size='8pt')
        self.report_plot.showGrid(x=True, y=True, alpha=0.12)

        # Y 軸：顯示 mm 並以小數精度呈現
        left_axis = self.report_plot.getAxis('left')
        left_axis.setTextPen(C['text_lo'])
        left_axis.enableAutoSIPrefix(False)  # 不要自動縮放成 μ/m/k 等

        # X 軸：使用自訂 tick formatter
        self._x_axis = self.report_plot.getAxis('bottom')
        self._x_axis.setTextPen(C['text_lo'])
        self._x_axis.setStyle(tickTextOffset=4)

        self.report_curves: Dict[int, pg.PlotDataItem] = {}
        for i in range(6):
            c = self.report_plot.plot(
                pen=pg.mkPen(color=CH_COLORS[i], width=1.5),
            )
            self.report_curves[i] = c
        chart_col.addWidget(self.report_plot, stretch=1)

        # ── 圖表下方控制列：模式切換（左）＋ 自製圖例（右）
        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(4)

        # 模式切換
        self.rb_time_mm = QRadioButton("time-mm-°C")
        self.rb_temp_mm = QRadioButton("°C-mm")
        self.rb_time_mm.setChecked(True)
        bg = QButtonGroup(self)
        bg.addButton(self.rb_time_mm)
        bg.addButton(self.rb_temp_mm)
        self.rb_time_mm.toggled.connect(self._refresh_report_plot)
        bottom_row.addWidget(self.rb_time_mm)
        bottom_row.addWidget(self.rb_temp_mm)

        bottom_row.addStretch()

        # 自製圖例：色線 + 文字，水平排列
        ch_names = ['CH1', 'CH2', 'CH3', 'CH4', 'CH5', 'CH6']
        for i, (color, name) in enumerate(zip(CH_COLORS, ch_names)):
            dot = QLabel("━")
            dot.setStyleSheet(f"color:{color}; font-size:14px; background:transparent;")
            dot.setFixedWidth(18)
            bottom_row.addWidget(dot)
            lbl = QLabel(f"{name} 變形(mm)")
            lbl.setStyleSheet(f"""
                color:{C['text_mid']};
                font-family:'Consolas','Courier New',monospace;
                font-size:9px;
                background:transparent;
            """)
            bottom_row.addWidget(lbl)
            if i < 5:
                bottom_row.addSpacing(6)

        chart_col.addLayout(bottom_row)

        upper.addLayout(chart_col, stretch=65)
        lay.addLayout(upper, stretch=45)

        # ── 下半：測試資料表格
        lay.addWidget(_label("測試資料", C['text_mid'], size=9, spacing=1))

        self.tbl_report = QTableWidget(0, 11)
        self.tbl_report.setHorizontalHeaderLabels([
            "測試日期", "測試時間", "流水批號", "材料名稱", "測試方法", "組別",
            "寬度 mm", "深度 mm", "跨距 mm", "變形 mm", "負載"
        ])
        for col in range(11):
            self.tbl_report.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.ResizeToContents
            )
        self.tbl_report.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_report.verticalHeader().setVisible(False)
        lay.addWidget(self.tbl_report, stretch=35)

        # ── 匯出按鈕
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_export = QPushButton("匯出報告")
        btn_export.setStyleSheet(_btn_style(C['amber']))
        btn_export.clicked.connect(self._export_report)
        btn_row.addWidget(btn_export)
        lay.addLayout(btn_row)

        return frame

    # ─────────────────────────────────────────────
    #  記錄管理
    # ─────────────────────────────────────────────
    def add_test_record(self, record: TestRecord):
        """外部（monitor_panel）停止測試後呼叫此方法，新增一筆記錄"""
        self.records.append(record)
        self._save_record_to_disk(record)
        self._refresh_record_table()

    def _refresh_record_table(self):
        self.tbl_records.setRowCount(0)
        for idx, rec in enumerate(self.records):
            r = self.tbl_records.rowCount()
            self.tbl_records.insertRow(r)
            self.tbl_records.setItem(r, 0, QTableWidgetItem(str(idx + 1)))
            self.tbl_records.setItem(r, 1, QTableWidgetItem(rec.name or rec.test_name))

    def _on_record_selected(self):
        rows = self.tbl_records.selectedItems()
        if not rows:
            return
        row = self.tbl_records.currentRow()
        if row < 0 or row >= len(self.records):
            return
        rec = self.records[row]
        self._selected_record = rec

        self.lbl_i_test_name.setText(rec.test_name)
        self.lbl_i_customer.setText(rec.customer)
        self.lbl_i_serial.setText(rec.serial_no)
        self.lbl_i_material.setText(rec.material)
        self.lbl_i_date.setText(rec.test_date)

        # 更新通道資料表
        self.tbl_ch_data.setRowCount(0)
        for ch in rec.channels:
            r = self.tbl_ch_data.rowCount()
            self.tbl_ch_data.insertRow(r)
            self.tbl_ch_data.setItem(r, 0, QTableWidgetItem(rec.customer))
            self.tbl_ch_data.setItem(r, 1, QTableWidgetItem(rec.material))
            self.tbl_ch_data.setItem(r, 2, QTableWidgetItem(rec.serial_no))
            self.tbl_ch_data.setItem(r, 3, QTableWidgetItem(ch.get("group", "--")))
            self.tbl_ch_data.setItem(r, 4, QTableWidgetItem(rec.test_date))
            self.tbl_ch_data.setItem(r, 5, QTableWidgetItem(rec.test_time))
            self.tbl_ch_data.setItem(r, 6, QTableWidgetItem(rec.test_method))

    def _import_to_report(self):
        if not self._selected_record:
            QMessageBox.information(self, "提示", "請先選擇一筆測試記錄。")
            return
        rec = self._selected_record
        self._report_record = rec

        self.le_r_test_name.setText(rec.test_name)
        self.le_r_customer.setText(rec.customer)
        self.le_r_medium.setText(rec.medium)
        self.le_r_serial.setText(rec.serial_no)
        self.le_r_material.setText(rec.material)
        self.le_r_method.setText(rec.test_method)
        self.le_r_date.setText(rec.test_date)
        self.le_r_addr.setText(rec.client_addr)

        # 更新右側表格
        self.tbl_report.setRowCount(0)
        for i, ch in enumerate(rec.channels):
            r = self.tbl_report.rowCount()
            self.tbl_report.insertRow(r)
            self.tbl_report.setItem(r, 0, QTableWidgetItem(rec.test_date))
            self.tbl_report.setItem(r, 1, QTableWidgetItem(rec.test_time))
            self.tbl_report.setItem(r, 2, QTableWidgetItem(rec.serial_no))
            self.tbl_report.setItem(r, 3, QTableWidgetItem(rec.material))
            self.tbl_report.setItem(r, 4, QTableWidgetItem(rec.test_method))
            self.tbl_report.setItem(r, 5, QTableWidgetItem(ch.get("group", f"CH{i+1}")))
            self.tbl_report.setItem(r, 6, QTableWidgetItem(str(ch.get("width", "--"))))
            self.tbl_report.setItem(r, 7, QTableWidgetItem(str(ch.get("depth", "--"))))
            self.tbl_report.setItem(r, 8, QTableWidgetItem(str(ch.get("span", "--"))))
            self.tbl_report.setItem(r, 9, QTableWidgetItem(str(ch.get("deflection", "--"))))
            self.tbl_report.setItem(r, 10, QTableWidgetItem(str(ch.get("load", "--"))))

        self._refresh_report_plot()

    def _import_to_viewer(self):
        """匯入至檢視器（簡化：僅在圖表顯示）"""
        self._import_to_report()

    @staticmethod
    def _make_time_ticks(total_seconds: float):
        """
        根據總時間長度決定 X 軸刻度格式與間隔。
        回傳 (ticks, label_fn)：
          ticks   — list[float]，刻度位置（秒）
          label_fn — callable(float) -> str，格式化標籤
        """
        if total_seconds <= 0:
            return [], str

        # 決定刻度間隔（秒）
        targets = [1, 2, 5, 10, 15, 20, 30,
                   60, 120, 300, 600, 900, 1800,
                   3600, 7200, 10800]
        # 希望大約 6~10 個刻度
        interval = targets[-1]
        for t in targets:
            if total_seconds / t <= 10:
                interval = t
                break

        ticks = []
        v = 0.0
        while v <= total_seconds + 1e-6:
            ticks.append(v)
            v += interval

        if total_seconds < 60:
            # 顯示秒
            label_fn = lambda s: f"{int(round(s))}s"
        elif total_seconds < 3600:
            # 顯示 分:秒
            def label_fn(s):
                m = int(s) // 60
                sec = int(s) % 60
                return f"{m}:{sec:02d}"
        else:
            # 顯示 時:分
            def label_fn(s):
                h = int(s) // 3600
                m = (int(s) % 3600) // 60
                return f"{h}:{m:02d}"

        return ticks, label_fn

    def _refresh_report_plot(self):
        if not self._report_record:
            return
        rec = self._report_record
        use_time = self.rb_time_mm.isChecked()

        all_x: List[float] = []
        for i in range(6):
            defl = rec.deflection_data.get(i, [])
            if use_time:
                xs = rec.time_data
                ys = defl
            else:
                xs = rec.temp_data.get(i, [])
                ys = defl

            min_len = min(len(xs), len(ys))
            if min_len > 0:
                self.report_curves[i].setData(xs[:min_len], ys[:min_len])
                all_x.extend(xs[:min_len])
            else:
                self.report_curves[i].setData([], [])

        # ── X 軸範圍 & 刻度（只對 time 模式套用智慧刻度）
        if use_time and all_x:
            t_max = max(all_x)
            t_min = min(all_x)
            span  = t_max - t_min if t_max > t_min else t_max

            ticks, label_fn = self._make_time_ticks(span if t_min == 0 else t_max)

            # 以實際資料起訖為視野，不多留空白
            self.report_plot.setXRange(t_min, t_max, padding=0.02)

            # 套用自訂刻度標籤
            tick_strings = [(v, label_fn(v)) for v in ticks]
            self._x_axis.setTicks([tick_strings, []])
        else:
            # °C 模式：恢復自動刻度
            self.report_plot.enableAutoRange(axis='x')
            self._x_axis.setTicks(None)

    def _remove_record(self):
        row = self.tbl_records.currentRow()
        if row < 0 or row >= len(self.records):
            return
        name = self.records[row].name
        reply = QMessageBox.question(
            self, "確認移除", f"確定移除「{name}」？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            del self.records[row]
            self._refresh_record_table()

    def _export_record(self):
        row = self.tbl_records.currentRow()
        if row < 0 or row >= len(self.records):
            QMessageBox.information(self, "提示", "請先選擇一筆測試記錄。")
            return
        rec = self.records[row]
        default = f"{rec.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        path, _ = QFileDialog.getSaveFileName(self, "匯出測試記錄", default, "JSON Files (*.json)")
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(rec.to_dict(), f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "完成", f"已匯出：\n{path}")

    def _export_report(self):
        if not self._report_record:
            QMessageBox.information(self, "提示", "請先匯入一筆測試記錄至報告產生器。")
            return
        rec = self._report_record

        # ── 檢查 openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, Reference, ScatterChart, Series
            from openpyxl.chart.series import SeriesLabel
        except ImportError:
            QMessageBox.critical(
                self, "缺少套件",
                "請先安裝 openpyxl：\n\npip install openpyxl\n\n安裝後重新啟動程式。"
            )
            return

        default = f"Report_{rec.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "匯出報告", default, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # ── 全局背景黑（前100列 x 20欄）
        from openpyxl.styles import PatternFill as _PF
        _bg = _PF("solid", fgColor="080A0D")
        for _r in range(1, 101):
            for _c in range(1, 42):
                ws.cell(_r, _c).fill = _bg
        ws.sheet_properties.tabColor = "FFC233"


        # ── 樣式定義
        BG          = "080A0D"   # 全局背景黑
        AMBER       = "FFC233"   # 琥珀金邊框
        hdr_font   = Font(name="Consolas", bold=True, color="FFC233", size=10)
        hdr_fill   = PatternFill("solid", fgColor=BG)
        title_font = Font(name="Consolas", bold=True, color="FFB300", size=13)
        key_font   = Font(name="Consolas", bold=True, color="8892A4", size=10)
        val_font   = Font(name="Consolas", color="E8EAF0", size=10)
        tbl_hdr_fill = PatternFill("solid", fgColor=BG)
        tbl_row_fill = PatternFill("solid", fgColor=BG)
        bg_fill      = PatternFill("solid", fgColor=BG)
        amber_side   = Side(style="thin", color=AMBER)
        thin_dark    = Side(style="thin", color="1A1F2A")
        border       = Border(left=amber_side, right=amber_side, top=amber_side, bottom=amber_side)
        center = Alignment(horizontal="center", vertical="center")

        def _set(row, col, value, font=None, fill=None, align=None, bdr=None):
            cell = ws.cell(row=row, column=col, value=value)
            if font:  cell.font   = font
            if fill:  cell.fill   = fill
            if align: cell.alignment = align
            if bdr:   cell.border = bdr
            return cell

        # ── 報告標題（列 1）
        ws.merge_cells("A1:K1")
        _set(1, 1,
             self.le_report_title.text() or "HDT/VICAT Test Report",
             font=title_font,
             fill=PatternFill("solid", fgColor="080A0D"),
             align=Alignment(horizontal="center", vertical="center"),
             bdr=border)
        # 合併欄 B1:K1 補上連續邊框
        for _c in range(2, 12):
            ws.cell(1, _c).border = Border(
                right=amber_side if _c == 11 else None,
                top=amber_side, bottom=amber_side
            )
        ws.row_dimensions[1].height = 28

        # ── 基本資訊（列 2-10）
        info_rows = [
            ("測試名稱", self.le_r_test_name.text()),
            ("客戶名稱", self.le_r_customer.text()),
            ("使用介質", self.le_r_medium.text()),
            ("流水批號", self.le_r_serial.text()),
            ("材料名稱", self.le_r_material.text()),
            ("測試方法", self.le_r_method.text()),
            ("測試日期", self.le_r_date.text()),
            ("客戶地址", self.le_r_addr.text()),
        ]
        for i, (k, v) in enumerate(info_rows, start=2):
            _set(i, 1, k, font=key_font, fill=hdr_fill, align=center, bdr=border)
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
            _set(i, 2, v, font=val_font, fill=PatternFill("solid", fgColor="080A0D"),
                 align=Alignment(vertical="center"), bdr=border)
            # 合併欄的右側邊框：C、D 欄補上 amber 右邊框，保持框線連續
            amber_r = Border(right=amber_side, top=amber_side, bottom=amber_side)
            ws.cell(i, 3).border = amber_r
            ws.cell(i, 4).border = Border(right=amber_side, top=amber_side, bottom=amber_side)
            ws.row_dimensions[i].height = 18

        # ── 空行
        blank_row = 11

        # ── 通道測試資料表格標頭（列 12）
        # 對應圖片欄位：編號 寬度 深度 跨距 軟化點 應力 軟化溫度 速率
        tbl_start = blank_row + 1
        tbl_headers = ["編號", "寬度", "深度", "跨距", "軟化點", "應力", "軟化溫度", "速率"]
        for col, h in enumerate(tbl_headers, start=1):
            _set(tbl_start, col, h,
                 font=hdr_font, fill=tbl_hdr_fill, align=center, bdr=border)
        ws.row_dimensions[tbl_start].height = 18

        # ── 通道資料列
        for i, ch in enumerate(rec.channels):
            r = tbl_start + 1 + i
            row_vals = [
                i + 1,                             # 編號
                ch.get("width",      "--"),        # 寬度
                ch.get("depth",      "--"),        # 深度
                ch.get("span",       "--"),        # 跨距
                ch.get("deflection", "--"),        # 軟化點（變形量）
                ch.get("load",       "--"),        # 應力（負載）
                ch.get("soft_temp",  "--"),        # 軟化溫度
                ch.get("rate",       "--"),        # 速率
            ]
            for col, val in enumerate(row_vals, start=1):
                _set(r, col, val, font=val_font, fill=tbl_row_fill,
                     align=center, bdr=border)
            ws.row_dimensions[r].height = 16

        data_end_row = tbl_start + len(rec.channels)

        # ── 時間序列資料（獨立工作表，供圖表參照）
        ws2 = wb.create_sheet("ChartData")
        if rec.time_data:
            # 標頭
            active_ch = []
            ws2.cell(1, 1, "時間(s)")
            for i in range(6):
                if rec.deflection_data.get(i):
                    ws2.cell(1, 2 + len(active_ch), f"CH{i+1} 變形(mm)")
                    active_ch.append(i)

            # 資料
            for idx, t in enumerate(rec.time_data):
                r2 = idx + 2
                ws2.cell(r2, 1, round(t, 2))
                for col_offset, ch_i in enumerate(active_ch):
                    defl = rec.deflection_data.get(ch_i, [])
                    ws2.cell(r2, 2 + col_offset,
                             round(defl[idx], 4) if idx < len(defl) else None)

            data_rows = len(rec.time_data)

            # ── X 軸時間長度，決定格式
            t_total = rec.time_data[-1] if rec.time_data else 0
            if t_total < 60:
                x_axis_title = "時間 (s)"
                x_num_fmt    = '0.0"s"'
            elif t_total < 3600:
                x_axis_title = "時間 (s)"
                x_num_fmt    = '0"s"'
            else:
                x_axis_title = "時間 (s)"
                x_num_fmt    = '0"s"'

            # ── 使用 ScatterChart（折線模式）：X/Y 軸都能顯示真實數值
            from openpyxl.chart import ScatterChart, Series

            chart = ScatterChart()
            chart.title   = None
            chart.style   = 10
            chart.height  = 16
            chart.width   = 28

            # Y 軸：mm，小數三位，強制顯示刻度數字
            chart.y_axis.title        = None
            chart.y_axis.numFmt       = '0.000'
            chart.y_axis.crosses      = "min"
            chart.y_axis.delete       = False
            chart.y_axis.axId         = 10
            chart.y_axis.tickLblPos   = "low"
            chart.y_axis.txPr         = None   # 讓 XML 修補處理旋轉

            # X 軸：秒，緊貼資料不留空白，強制顯示刻度數字
            # X 軸：秒，緊貼資料不留空白，強制顯示刻度數字
            chart.x_axis.title         = None
            chart.x_axis.crosses       = "autoZero"
            chart.x_axis.delete        = False
            chart.x_axis.axId          = 20
            chart.x_axis.tickLblPos    = "low"   # 標籤在底部
            chart.x_axis.scaling.min   = 0
            chart.x_axis.scaling.max   = float(round(t_total, 3))

            CHART_COLORS = ["FF6B6B", "FFA94D", "69DB7C", "4FC3F7", "DA77F2", "F783AC"]

            x_ref = Reference(ws2, min_col=1, min_row=2, max_row=data_rows + 1)

            for col_offset, ch_i in enumerate(active_ch):
                y_ref = Reference(ws2,
                                  min_col=2 + col_offset,
                                  min_row=1,           # 含標頭，讓 series 自動取名
                                  max_row=data_rows + 1)
                ser = Series(y_ref, x_ref, title_from_data=True)
                # 折線樣式（無標記點）
                ser.graphicalProperties.line.solidFill = CHART_COLORS[ch_i % len(CHART_COLORS)]
                ser.graphicalProperties.line.width = 18000   # 1.8 pt in EMU/100
                ser.marker.symbol = "none"
                chart.series.append(ser)

            # 圖表放在 Report 工作表，通道表格下方留兩列空白
            # 標題寫在儲存格，圖表框內不顯示標題（避免卡格線）
            chart_anchor_row = data_end_row + 3
            title_cell = ws.cell(chart_anchor_row, 1, "變形 vs 時間")
            title_cell.font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
            ws.add_chart(chart, f"A{chart_anchor_row + 1}")

        # ── 欄寬自動調整（Report 表）
        col_widths = [12, 10, 14, 14, 14, 8, 10, 10, 10, 10, 8]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        # ── 儲存
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.critical(self, "錯誤", f"無法寫入檔案，請確認檔案未被其他程式開啟：\n{path}")
            return

        # ── 修補圖表 XML：plotArea 手動佈局 + Y 軸標題橫排
        try:
            import zipfile, shutil, re
            tmp_path = path + ".tmp"

            # openpyxl 實際產生的 bodyPr 標籤（無任何屬性，帶 xmlns）：
            #   <a:bodyPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"/>
            # 加上 rot="0" vert="horz" 即可讓標題橫排
            _XMLNS_A = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
            _BODYPR_ORIG  = f'<a:bodyPr {_XMLNS_A}/>'
            _BODYPR_HORIZ = f'<a:bodyPr {_XMLNS_A} rot="0" vert="horz"/>'

            def _patch_axis_title_horiz(valax_xml: str) -> str:
                """將 <valAx> 區段內 <title>...</title> 裡的 bodyPr 改為橫排。"""
                m_title = re.search(r'<title>.*?</title>', valax_xml, re.DOTALL)
                if not m_title:
                    return valax_xml
                title_fixed = m_title.group(0).replace(_BODYPR_ORIG, _BODYPR_HORIZ)
                return valax_xml[:m_title.start()] + title_fixed + valax_xml[m_title.end():]

            with zipfile.ZipFile(path, 'r') as zin, \
                 zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if re.match(r'xl/charts/chart\d+\.xml', item.filename):
                        xml = data.decode('utf-8')

                        # ── 1. 插入 plotArea 手動佈局
                        #    x=0.25 → 圖表左邊界從 25% 開始，給 Y 軸橫排標題 + 數字足夠空間
                        #    h=0.80 → 圖表底部留 20%，X 軸標題「時間(s)」不被截斷
                        manual_layout = (
                            '<c:layout>'
                            '<c:manualLayout>'
                            '<c:layoutTarget val="inner"/>'
                            '<c:xMode val="edge"/>'
                            '<c:yMode val="edge"/>'
                            '<c:x val="0.10"/>'
                            '<c:y val="0.10"/>'
                            '<c:w val="0.72"/>'
                            '<c:h val="0.80"/>'
                        )
                        if '<c:layout/>' in xml:
                            xml = xml.replace('<c:layout/>', manual_layout)
                        elif '<c:layout>' not in xml:
                            xml = xml.replace('<c:plotArea>', '<c:plotArea>' + manual_layout, 1)

                        # ── 2. 深色主題：背景 + 文字白色 + 移除外框白邊
                        _CHART_BG = "0E1117"
                        _NS_A = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
                        _WHITE = "FFFFFF"

                        # chartSpace 整體背景 + 移除外框白邊（合併成一個 spPr）
                        _chartspace_spPr = (
                            f'<spPr>'
                            f'<a:solidFill {_NS_A}><a:srgbClr val="{_CHART_BG}"/></a:solidFill>'
                            f'<a:ln {_NS_A}><a:noFill/></a:ln>'
                            f'</spPr>'
                        )
                        xml = xml.replace('<chart>', _chartspace_spPr + '<chart>', 1)

                        # 繪圖區背景
                        _plot_spPr = (
                            f'<spPr>'
                            f'<a:solidFill {_NS_A}><a:srgbClr val="{_CHART_BG}"/></a:solidFill>'
                            f'</spPr>'
                        )
                        xml = xml.replace('</plotArea>', _plot_spPr + '</plotArea>', 1)

                        # 白色文字 txPr（圖例 + 軸刻度共用）
                        _white_txPr = (
                            f'<txPr><a:bodyPr {_NS_A}/>'
                            f'<a:lstStyle {_NS_A}/>'
                            f'<a:p {_NS_A}><a:pPr><a:defRPr b="0">'
                            f'<a:solidFill><a:srgbClr val="{_WHITE}"/></a:solidFill>'
                            f'</a:defRPr></a:pPr></a:p></txPr>'
                        )

                        # 圖例：白色文字 + 深色背景 + 無邊框
                        xml = xml.replace(
                            '<legend><legendPos val="r"/></legend>',
                            '<legend><legendPos val="r"/>' + _white_txPr +
                            f'<spPr><a:solidFill {_NS_A}><a:srgbClr val="{_CHART_BG}"/></a:solidFill>'
                            f'<a:ln {_NS_A}><a:noFill/></a:ln></spPr></legend>'
                        )

                        # 圖表主標題：白色文字（只替換第一個 defRPr/）
                        xml = xml.replace(
                            '<a:defRPr/></a:pPr>',
                            f'<a:defRPr><a:solidFill {_NS_A}>'
                            f'<a:srgbClr val="{_WHITE}"/></a:solidFill></a:defRPr></a:pPr>',
                            1
                        )

                        # 軸刻度數字：白色（每個 valAx 結尾前插入 txPr）
                        xml = xml.replace('</valAx>', _white_txPr + '</valAx>')

                        # ── 3. Y 軸 title 橫排
                        #    openpyxl ScatterChart 產生兩個 <valAx>：
                        #      第一個 axId=10 → 時間(X軸)
                        #      第二個 axId=20 → 變形(Y軸) ← 這個才需要橫排
                        #    以 title 內文「變形」來定位，避免順序依賴
                        for m_ax in re.finditer(r'<valAx\b.*?</valAx>', xml, re.DOTALL):
                            ax_xml = m_ax.group(0)
                            t_match = re.search(r'<a:t>(.*?)</a:t>', ax_xml)
                            if t_match and '變形' in t_match.group(1):
                                ax_fixed = _patch_axis_title_horiz(ax_xml)
                                xml = xml[:m_ax.start()] + ax_fixed + xml[m_ax.end():]
                                break

                        data = xml.encode('utf-8')
                    zout.writestr(item, data)
            shutil.move(tmp_path, path)
        except Exception as e:
            print(f"[REPORT] XML 修補失敗（圖表仍可用）: {e}")

        QMessageBox.information(self, "完成", f"報告已匯出：\n{path}")

    # ─────────────────────────────────────────────
    #  磁碟持久化（簡易 JSON）
    # ─────────────────────────────────────────────
    def _save_record_to_disk(self, rec: TestRecord):
        try:
            fp = DATA_DIR / f"{rec.record_id or rec.name}.json"
            with open(fp, 'w', encoding='utf-8') as f:
                # 只儲存摘要資料（折線圖資料量大，選擇性儲存）
                d = rec.to_dict()
                json.dump(d, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[REPORT] 無法儲存記錄: {e}")

    def _load_records_from_disk(self):
        """程式啟動時清空上次的測試記錄（json 檔），避免殘留舊資料。"""
        try:
            for fp in DATA_DIR.glob("*.json"):
                try:
                    fp.unlink()
                except Exception as e:
                    print(f"[REPORT] 無法刪除舊記錄 {fp.name}: {e}")
        except Exception as e:
            print(f"[REPORT] 清空記錄目錄時發生錯誤: {e}")
        # records 列表已在 __init__ 初始化為 []，無需額外清空
        self._refresh_record_table()